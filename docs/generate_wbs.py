"""Generate Work Breakdown Structure (WBS) Excel for AuctionHub.

One sheet, single table, columns: Level | WBS | Task Description | Notes.
Hierarchy follows Fernando's combined Buyer/Seller WBS.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"c:\dev\FYP-26-S2-24\docs\AuctionHub_WBS.xlsx"

# (level, wbs_code, description, notes)
ROWS = [
    # 1. Functional Hierarchy (the product itself)
    (1, "1", "Functional Hierarchy", "What the product does (per assessor's combined Buyer/Seller model)"),

    # 1.1 Unregistered User (Guest) — read-only browsing + authentication entry points
    (2, "1.1", "Unregistered User (Guest)", "Read-only browsing + sign up / login; no interactive auction actions (no Bid / Watchlist / Q&A)"),
    (3, "1.1.1", "Browse homepage / Popular Right Now (trending)", "Read-only landing page with featured & trending listings"),
    (3, "1.1.2", "Search auctions (keyword / category / filters / sort)", "Read-only search; no bidding entry points exposed"),
    (3, "1.1.3", "View auction detail page", "View images, description, current price (read-only; no Bid button)"),
    (3, "1.1.4", "View public bid history", "Read-only, anonymised bidder names"),
    (3, "1.1.5", "View public seller profile", "Ratings, reviews, active listings (read-only)"),
    (3, "1.1.6", "Sign up (Register new account)", "Email + password; routes user to Registered User flows after success"),
    (3, "1.1.7", "Login", "Existing-user entry point; Bearer token issued (per-tab session)"),
    (3, "1.1.8", "Forgot / Reset password (OTP via email)", "OTP store with expiry"),

    # 1.2 Registered User (Buyer + Seller combined)
    (2, "1.2", "Registered User (Buyer + Seller - any account can do both)", "Combined role per assessor feedback"),

    # 1.2.1 Account & Authentication
    (3, "1.2.1", "Account & Authentication", ""),
    (4, "1.2.1.1", "Two-factor authentication (TOTP)", "Optional second factor"),
    (4, "1.2.1.2", "View profile", "Public + private fields"),
    (4, "1.2.1.3", "Edit profile (name, contact, address)", ""),
    (4, "1.2.1.4", "Upload / change profile photo", "Image validation"),
    (4, "1.2.1.5", "Change password", "Requires current password"),
    (4, "1.2.1.6", "Account settings", "Notification & privacy preferences"),
    (4, "1.2.1.7", "Delete account (PDPA anonymisation)", "Anonymise instead of hard delete"),
    (4, "1.2.1.8", "Logout", "Invalidates token"),

    # 1.2.2 Buying
    (3, "1.2.2", "Buying", ""),
    (4, "1.2.2.1", "Place manual bid", "SELECT ... FOR UPDATE concurrency"),
    (4, "1.2.2.2", "Set / cancel auto-bid (proxy)", "AES-GCM encrypted max amount"),
    (4, "1.2.2.3", "View / edit active auto-bid", "Per-buyer increment"),
    (4, "1.2.2.4", "Add / remove watchlist item", ""),
    (4, "1.2.2.5", "View watchlist", ""),
    (4, "1.2.2.6", "View bidding history", "Win / loss status"),
    (4, "1.2.2.7", "Ask question on a listing", "Q&A"),
    (4, "1.2.2.8", "Rate seller after order", "Score + comment"),
    (4, "1.2.2.9", "Report user or listing", "Moderation queue"),
    (4, "1.2.2.10", "Personalised recommendations", "Collaborative filtering + trending fallback"),
    (4, "1.2.2.11", "View orders", "Pending / paid / shipped / delivered"),
    (4, "1.2.2.12", "Pay for won auction", "Payment method selection"),
    (4, "1.2.2.13", "Confirm receipt of delivered item", ""),
    (4, "1.2.2.14", "Order messaging with seller", ""),
    (4, "1.2.2.15", "Request refund", ""),

    # 1.2.3 Selling
    (3, "1.2.3", "Selling (requires SELLER role)", ""),
    (4, "1.2.3.1", "Create new auction listing", "Standard / Dutch / Blind types"),
    (4, "1.2.3.2", "Upload listing images", ""),
    (4, "1.2.3.3", "Edit existing auction", "Pre-bid only"),
    (4, "1.2.3.4", "Cancel auction", ""),
    (4, "1.2.3.5", "Seller dashboard (active / ended)", ""),
    (4, "1.2.3.6", "Seller analytics (revenue, bids)", ""),
    (4, "1.2.3.7", "Featured / promoted listings", ""),
    (4, "1.2.3.8", "Reply to buyer questions", ""),
    (4, "1.2.3.9", "Rate buyer after order", ""),
    (4, "1.2.3.10", "Declare auction winner", "Early or post-expiry"),
    (4, "1.2.3.11", "Advance shipping status", "Paid -> Shipped -> Delivered"),
    (4, "1.2.3.12", "Approve / reject refund request", ""),
    (4, "1.2.3.13", "Public seller profile", "Visible to all users"),
    (4, "1.2.3.14", "Order messaging with buyer", ""),

    # 1.2.4 Shared
    (3, "1.2.4", "Shared features", ""),
    (4, "1.2.4.1", "Notifications (in-app)", "Outbid, won, order updates"),
    (4, "1.2.4.2", "Mark notification as read", ""),
    (4, "1.2.4.3", "Support chat with admin", ""),

    # 1.3 Admin
    (2, "1.3", "Admin", "Platform operator"),
    (3, "1.3.1", "Admin dashboard (metrics, activity)", "KPIs"),
    (3, "1.3.2", "User moderation", ""),
    (4, "1.3.2.1", "View user list", ""),
    (4, "1.3.2.2", "Ban user", ""),
    (4, "1.3.2.3", "Unban user", ""),
    (3, "1.3.3", "Listing moderation", ""),
    (4, "1.3.3.1", "View flagged listings", ""),
    (4, "1.3.3.2", "Remove / hide listing", ""),
    (4, "1.3.3.3", "Restore listing", ""),
    (3, "1.3.4", "Category management (CRUD)", ""),
    (4, "1.3.4.1", "Create category", ""),
    (4, "1.3.4.2", "Edit category", ""),
    (4, "1.3.4.3", "Soft-delete category", ""),
    (4, "1.3.4.4", "Restore category", ""),
    (3, "1.3.5", "Order management", ""),
    (3, "1.3.6", "Reports & analytics", ""),
    (4, "1.3.6.1", "Resolve user / listing reports", ""),
    (4, "1.3.6.2", "Generate / export analytics reports", "CSV / PDF export"),
    (3, "1.3.7", "Support chat (admin side)", ""),
    (3, "1.3.8", "Database / system view", "Health / backup status"),

    # 1.4 Cross-cutting (technical scope realised in prototype)
    (2, "1.4", "Cross-cutting / Platform features", "Realised in prototype"),
    (3, "1.4.1", "Authentication & RBAC (Bearer token, per-tab sessions)", ""),
    (3, "1.4.2", "Auction lifecycle scheduler (PENDING -> ACTIVE -> ENDED)", "AuctionExpiryListener"),
    (3, "1.4.3", "Real-time updates (SSE backend + polling fallback)", "Cloudflare workaround"),
    (3, "1.4.4", "Security (CORS, CSP, AES-GCM encryption, input validation)", ""),
    (3, "1.4.5", "Database (PostgreSQL + HikariCP pool, migrations)", ""),
    (3, "1.4.6", "Deployment (Render hosted demo)", "Live URL"),

    # 2. Project Management
    (1, "2", "Project Management", "Team coordination & delivery process"),
    (2, "2.1", "Team meetings & stand-ups", "Weekly sync"),
    (2, "2.2", "Sprint planning & retrospectives", "Sprint 1-4 retrospectives"),
    (2, "2.3", "Jira backlog & Scrum board", "User stories, tickets, status"),
    (2, "2.4", "Workload distribution & peer assessment", "Per submission"),
    (2, "2.5", "Risk register & mitigation", "Identified risks tracked"),
    (2, "2.6", "Progress reporting to supervisor", "Bi-weekly check-ins"),

    # 3. Testing & Quality Assurance
    (1, "3", "Testing & Quality Assurance", "Verification of all features"),
    (2, "3.1", "Unit testing (JUnit + Mockito)", "Per DAO/Servlet"),
    (2, "3.2", "Integration testing", "End-to-end flows"),
    (2, "3.3", "Concurrency testing", "FOR UPDATE, auto-bid races"),
    (2, "3.4", "Security testing", "RBAC, IDOR, input validation"),
    (2, "3.5", "User Acceptance Testing (UAT)", "Demo accounts walkthrough"),
    (2, "3.6", "Code review", "Peer review on PRs"),
    (2, "3.7", "Bug tracking & resolution", "SCRUM-### tickets"),

    # 4. Deployment & DevOps
    (1, "4", "Deployment & DevOps", "Hosting & operations"),
    (2, "4.1", "Local development environment", "Docker Compose, README"),
    (2, "4.2", "Database hosting (Supabase PostgreSQL)", "Schema + migrations"),
    (2, "4.3", "Application hosting (Render)", "Tomcat Docker image"),
    (2, "4.4", "Frontend hosting (React build)", "Vite production build"),
    (2, "4.5", "CI/CD pipeline (GitHub Actions)", "Build, test, deploy"),
    (2, "4.6", "Database backup & restore strategy", "Manual + automated"),
    (2, "4.7", "Live demo URL & monitoring", "Render dashboard"),

    # 5. Documentation
    (1, "5", "Documentation", "Submission deliverables"),
    (2, "5.1", "PTM (Project Technical Manual)", "Architecture, DAO/Servlet specs"),
    (2, "5.2", "PUM (Project User Manual)", "Use cases, user journeys"),
    (2, "5.3", "PTD (Project Technical Documentation)", "Functional hierarchy, diagrams"),
    (2, "5.4", "ERD & Master Class Diagram", "Updated per sprint"),
    (2, "5.5", "Sequence diagrams (per feature)", "Auth, bid, auto-bid, order"),
    (2, "5.6", "API documentation", "Endpoints, request/response"),
    (2, "5.7", "Presentation slides (Submission 1-3)", "Mid + final demos"),
    (2, "5.8", "Peer assessment forms", "Per submission cycle"),
    (2, "5.9", "Record of Revision", "Document version control"),
    (2, "5.10", "README & setup guide", "Repo onboarding"),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WBS"

    # Title
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "AuctionHub - Work Breakdown Structure"
    title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill("solid", fgColor="1F4E78")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = "FYP 26-S2-24 | Single sheet | Hierarchy: Project -> Activity -> Task -> Subtask"
    sub.font = Font(italic=True, color="595959")
    sub.alignment = Alignment(horizontal="center")

    # Header row
    headers = ["Level", "WBS", "Task Description", "Notes"]
    header_row = 4
    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # Level colors (consistent with WBS templates)
    level_fill = {
        1: PatternFill("solid", fgColor="FCE4D6"),  # peach
        2: PatternFill("solid", fgColor="FFF2CC"),  # light yellow
        3: PatternFill("solid", fgColor="E2EFDA"),  # light green
        4: PatternFill("solid", fgColor="FFFFFF"),  # white
    }
    level_font = {
        1: Font(bold=True, size=12),
        2: Font(bold=True),
        3: Font(bold=True, color="375623"),
        4: Font(),
    }

    # Write data rows
    start_row = header_row + 1
    for i, (lvl, code, desc, note) in enumerate(ROWS):
        r = start_row + i
        ws.cell(row=r, column=1, value=lvl)
        ws.cell(row=r, column=2, value=code)
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=4, value=note)

        for col_idx in range(1, 5):
            c = ws.cell(row=r, column=col_idx)
            c.fill = level_fill.get(lvl, level_fill[4])
            c.font = level_font.get(lvl, level_font[4])
            c.border = border
            if col_idx in (1, 2):
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Indent task description by level for visual hierarchy
        indent_cell = ws.cell(row=r, column=3)
        indent_cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, indent=(lvl - 1) * 2
        )

    # Column widths
    widths = {1: 8, 2: 12, 3: 70, 4: 45}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header
    ws.freeze_panes = "A5"

    # Auto filter on header
    last_row = start_row + len(ROWS) - 1
    ws.auto_filter.ref = f"A{header_row}:D{last_row}"

    draw_diagram(ws)

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}  ({len(ROWS)} rows)")


def draw_diagram(ws) -> None:
    """Draw a horizontal tree diagram (Levels 1-3) to the right of the table."""
    DIAG_COL_START = 6   # column F
    DIAG_COL_END = 20    # column T (15 columns wide)
    LINE_COLOR = "595959"
    BOX_BORDER_COLOR = "1F4E78"

    # Column widths for diagram area
    for c in range(DIAG_COL_START, DIAG_COL_END + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11

    thin_line = Side(border_style="thin", color=LINE_COLOR)
    box_side = Side(border_style="medium", color=BOX_BORDER_COLOR)
    box_border = Border(left=box_side, right=box_side, top=box_side, bottom=box_side)

    def add_left_border(row: int, col: int) -> None:
        cell = ws.cell(row=row, column=col)
        old = cell.border
        cell.border = Border(
            left=thin_line, right=old.right, top=old.top, bottom=old.bottom
        )

    def add_bottom_border(row: int, col: int) -> None:
        cell = ws.cell(row=row, column=col)
        old = cell.border
        cell.border = Border(
            left=old.left, right=old.right, top=old.top, bottom=thin_line
        )

    def vline(from_row: int, to_row: int, col: int) -> None:
        for r in range(from_row, to_row + 1):
            add_left_border(r, col)

    def hline(row: int, from_col: int, to_col: int) -> None:
        for c in range(from_col, to_col + 1):
            add_bottom_border(row, c)

    def box(top: int, bottom: int, left: int, right: int, text: str,
            fill: str, font_color: str = "000000", bold: bool = True,
            size: int = 10) -> None:
        ws.merge_cells(start_row=top, start_column=left,
                       end_row=bottom, end_column=right)
        cell = ws.cell(row=top, column=left, value=text)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=bold, color=font_color, size=size)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        for r in range(top, bottom + 1):
            for c in range(left, right + 1):
                ws.cell(row=r, column=c).border = box_border

    # Diagram title (row 4 same as table header row)
    title_row = 4
    ws.merge_cells(start_row=title_row, start_column=DIAG_COL_START,
                   end_row=title_row, end_column=DIAG_COL_END)
    tcell = ws.cell(row=title_row, column=DIAG_COL_START,
                    value="Functional Hierarchy Diagram")
    tcell.font = Font(bold=True, color="FFFFFF", size=12)
    tcell.fill = PatternFill("solid", fgColor="2E75B6")
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    tcell.border = box_border

    # Layout (all boxes 4 cols wide so connector at boundary = box centre)
    # Guest:      cols 6-9   centre = boundary 7|8  -> col 8
    # Registered: cols 11-14 centre = boundary 12|13 -> col 13
    # Admin:      cols 16-19 centre = boundary 17|18 -> col 18
    # Root:       cols 11-14 (same as Registered)    -> col 13

    # Row plan
    ROOT_TOP, ROOT_BOTTOM = 6, 7
    TRUNK_DOWN_ROW = 8       # vertical line under root + horizontal trunk
    TRUNK_TO_L2_ROW = 9      # vertical lines down to each L2 box
    L2_TOP, L2_BOTTOM = 10, 11

    # Root box
    box(ROOT_TOP, ROOT_BOTTOM, 11, 14, "AuctionHub", "FCE4D6", size=13)

    # Vertical line down from root
    vline(TRUNK_DOWN_ROW, TRUNK_DOWN_ROW, 13)

    # Horizontal trunk across to all three branches
    hline(TRUNK_DOWN_ROW, 8, 18)

    # Vertical lines from trunk down to each L2 box
    vline(TRUNK_TO_L2_ROW, TRUNK_TO_L2_ROW, 8)
    vline(TRUNK_TO_L2_ROW, TRUNK_TO_L2_ROW, 13)
    vline(TRUNK_TO_L2_ROW, TRUNK_TO_L2_ROW, 18)

    # L2 boxes
    box(L2_TOP, L2_BOTTOM, 6, 9,
        "Unregistered User\n(Guest)", "FFF2CC", size=11)
    box(L2_TOP, L2_BOTTOM, 11, 14,
        "Registered User\n(Buyer + Seller)", "FFF2CC", size=11)
    box(L2_TOP, L2_BOTTOM, 16, 19, "Admin", "FFF2CC", size=11)

    # L3 + L4 tree under each L2 box
    # Each entry: (L3 title, [L4 sub-items])
    guest_tree = [
        ("1.1.1 Browse / Popular Right Now", []),
        ("1.1.2 Search (keyword/filters/sort)", []),
        ("1.1.3 View auction detail (no Bid)", []),
        ("1.1.4 View public bid history", []),
        ("1.1.5 View public seller profile", []),
        ("1.1.6 Sign up (Register)", []),
        ("1.1.7 Login", []),
        ("1.1.8 Forgot / Reset password (OTP)", []),
    ]
    reg_tree = [
        ("1.2.1 Account & Authentication", [
            "TOTP (2FA)",
            "View profile",
            "Edit profile",
            "Profile photo",
            "Change password",
            "Account settings",
            "Delete (PDPA anonymisation)",
            "Logout",
        ]),
        ("1.2.2 Buying", [
            "Place manual bid",
            "Set / cancel auto-bid",
            "View / edit active auto-bid",
            "Watchlist (add / remove / view)",
            "Bidding history",
            "Ask question",
            "Rate seller",
            "Report user / listing",
            "Personalised recommendations",
            "Orders (pay / confirm / refund)",
            "Order messaging",
        ]),
        ("1.2.3 Selling (SELLER role)", [
            "Create / Edit / Cancel listing",
            "Upload listing images",
            "Seller dashboard",
            "Seller analytics",
            "Featured / promoted listings",
            "Reply to questions",
            "Rate buyer",
            "Declare auction winner",
            "Advance shipping status",
            "Approve / reject refund",
            "Public seller profile",
            "Order messaging",
        ]),
        ("1.2.4 Shared", [
            "Notifications (in-app)",
            "Mark notification as read",
            "Support chat with admin",
        ]),
    ]
    admin_tree = [
        ("1.3.1 Dashboard", []),
        ("1.3.2 User moderation", [
            "View user list",
            "Ban user",
            "Unban user",
        ]),
        ("1.3.3 Listing moderation", [
            "View flagged listings",
            "Remove / hide listing",
            "Restore listing",
        ]),
        ("1.3.4 Category CRUD", [
            "Create category",
            "Edit category",
            "Soft-delete category",
            "Restore category",
        ]),
        ("1.3.5 Order management", []),
        ("1.3.6 Reports & Analytics", [
            "Resolve reports",
            "Export analytics (CSV/PDF)",
        ]),
        ("1.3.7 Support chat (Admin)", []),
        ("1.3.8 Database / system view", []),
    ]

    panel_side = Side(border_style="thin", color="1F4E78")
    panel_border = Border(left=panel_side, right=panel_side,
                          top=panel_side, bottom=panel_side)

    def draw_l4_panel(top: int, bottom: int, left: int, right: int,
                      items: list) -> None:
        ws.merge_cells(start_row=top, start_column=left,
                       end_row=bottom, end_column=right)
        text = "\n".join(f"  •  {it}" for it in items)
        cell = ws.cell(row=top, column=left, value=text)
        cell.fill = PatternFill("solid", fgColor="FFFFFF")
        cell.font = Font(size=8, color="262626")
        cell.alignment = Alignment(horizontal="left", vertical="top",
                                   wrap_text=True, indent=1)
        for r in range(top, bottom + 1):
            for c in range(left, right + 1):
                ws.cell(row=r, column=c).border = panel_border

    def draw_branch(items, left: int, right: int, centre: int,
                    parent_bottom: int) -> int:
        cursor = parent_bottom + 1
        for l3_text, l4_list in items:
            gap_row = cursor
            l3_top = gap_row + 1
            l3_bottom = l3_top + 1
            vline(gap_row, l3_top - 1, centre)
            box(l3_top, l3_bottom, left, right, l3_text, "E2EFDA",
                size=9, bold=True)
            if l4_list:
                panel_top = l3_bottom + 1
                panel_bottom = panel_top + len(l4_list) - 1
                draw_l4_panel(panel_top, panel_bottom, left, right, l4_list)
                for r in range(panel_top, panel_bottom + 1):
                    ws.row_dimensions[r].height = 14
                cursor = panel_bottom + 1
            else:
                cursor = l3_bottom + 1
        return cursor

    draw_branch(guest_tree, 6, 9, 8, L2_BOTTOM)
    draw_branch(reg_tree, 11, 14, 13, L2_BOTTOM)
    draw_branch(admin_tree, 16, 19, 18, L2_BOTTOM)

    # Slightly taller rows for the root and L2 rows so titles breathe
    for r in (ROOT_TOP, ROOT_BOTTOM, L2_TOP, L2_BOTTOM):
        ws.row_dimensions[r].height = 22


if __name__ == "__main__":
    main()
